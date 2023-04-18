import cv2
from openpyxl import load_workbook
from functions import detect, detection_graph 

# Opening spreadsheet with number of latest image taken
wb = load_workbook('D:\scuola\MIT\cestino\programma\camera\saves.xlsx')
ws = wb.active
n = ws['B1'].value
# camera
camera = 2 # the number depends on wich camera is being used on the computer
vid = cv2.VideoCapture(camera) 
path =  "D:\scuola\MIT\cestino\programma\camera\images"
# Defining where each id of the labelmap needs to go
plastic = [5,6,8,22,25,28,30,37,40,41,42,43,45,46,48,49,50,56]
paper = [4,15,16,17,18,19,20,21,31,33,34,35,57]
generic = [23,24,26,29,32,38,44,47,51,52,58,59,60]
glass = [1,3,7,10,11,13,27]
not_accepted = [2,9,12,14,39,53,54,55]
on = True
#Preparing the main loop for when it will talk to an arduino
while on:
    print("press y to dispose")
    answer = input()
    if answer == "y":
        # Takes a picture of what is currently on the bin
        ret, frame = vid.read()
        img = cv2.imwrite(f'''{path}\image{n}.png''', frame)
        # object is the id from the labelmap of the object in picture
        materials = detect(detection_graph, f'''{path}\image{n}.png''')
        mat1 = int(materials[0])
        print(mat1)
        # moving the bin
        if mat1 in plastic:
            print("This is plastic")
            print(materials)
        elif mat1 in paper:
            print("This is paper")
            print(materials)
        elif mat1 in glass:
            print("This is glass")
            print(materials)
        elif mat1 in generic:
            print("This isn't recyclable")
            print(materials)
        else:
            print("I can't take this")
            print(materials)
        
        # updates the latest pic number to the saves file
        n += 1
        ws['B1'].value = n
        wb.save('D:\scuola\MIT\cestino\programma\camera\saves.xlsx')
    
    if answer == "close":
        on = False