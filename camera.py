import cv2
from openpyxl import load_workbook

# define a video capture object
camera = 2
vid = cv2.VideoCapture(camera) 
path =  "D:\scuola\MIT\cestino\programma\camera\images"
# open a saves file
wb = load_workbook('D:\scuola\MIT\cestino\programma\camera\saves.xlsx')
ws = wb.active
n = ws['B1'].value
snap_photo = False
while(True):
      
    # Capture the video frame
    # by frame
    ret, frame = vid.read()
  
    # Display the resulting frame
    cv2.imshow('frame', frame)
      
    # the 'q' button closes the program
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
    # the 's' button saves the frame as a png
    if snap_photo:
        img = cv2.imwrite(f'''{path}\image{n}.png''', frame)
        n += 1
        print(n)
        # updates the latest pic number to the saves file
        ws['B1'].value = n
        wb.save('saves.xlsx')
# After the loop release the cap object
vid.release()
# Destroy all the windows
cv2.destroyAllWindows()
